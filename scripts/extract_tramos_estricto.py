#!/usr/bin/env python3
"""Extract tramo reports with strict INICIO/FIN filter from Excel Planilla."""

from __future__ import annotations

import csv
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SOURCE_DIR = Path("/workspace/extracted/source/01 BASE DE DATOS")
OUTPUT_DIR = Path("/workspace/extracted/reports/tramos_filtrado_estricto")

TRAMOS = [
    ("125+300 - 125+500", "tramo_125+300_125+500", 125300, 125500),
    ("122+750 - 123+100", "tramo_122+750_123+100", 122750, 123100),
    ("146+500 - 146+520", "tramo_146+500_146+520", 146500, 146520),
    ("175+260 - 175+285", "tramo_175+260_175+285", 175260, 175285),
    ("214+580 - 214+630", "tramo_214+580_214+630", 214580, 214630),
    ("243+700 - 243+730", "tramo_243+700_243+730", 243700, 243730),
    ("282+280 - 282+400", "tramo_282+280_282+400", 282280, 282400),
    ("282+630 - 282+690", "tramo_282+630_282+690", 282630, 282690),
    ("284+750 - 284+790", "tramo_284+750_284+790", 284750, 284790),
    ("603+210 - 603+220", "tramo_603+210_603+220", 603210, 603220),
    ("605+585 - 605+600", "tramo_605+585_605+600", 605585, 605600),
]

COLUMNS = ["Descripción", "Unidad", "Fecha", "INICIO (KM)", "FIN (KM)", "METRADO"]


def to_num(value) -> int | None:
    if value is None:
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def format_progresiva_excel(value: int | None) -> str | None:
    """Match Excel TEXT(value, '000+000') used in INICIO/FIN columns."""
    if value is None:
        return None
    value = int(value)
    return f"{value // 1000:03d}+{value % 1000:03d}"


def serialize_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text if text else None


def serialize_metrado(value):
    if value is None:
        return None
    if isinstance(value, float):
        if value != value:
            return None
    return value


def strict_within_tramo(inicio: int, fin: int, r_start: int, r_end: int) -> bool:
    """Activity fully contained in tramo using INICIO (KM) and FIN (KM) only."""
    if inicio > fin:
        inicio, fin = fin, inicio
    return inicio >= r_start and fin <= r_end


def find_header_row(ws, max_search: int = 40) -> int | None:
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_search, values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if "INICIO (KM)" in vals and "FIN (KM)" in vals:
            return row_idx
    return None


def load_all_records() -> list[dict]:
    records: list[dict] = []
    for xlsx_path in sorted(SOURCE_DIR.glob("*.xlsx")):
        year = xlsx_path.stem
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb["Planilla"]
        header_row = find_header_row(ws)
        if not header_row:
            wb.close()
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if len(row) < 17:
                continue
            inicio_raw = to_num(row[14])
            fin_raw = to_num(row[15])
            if inicio_raw is None or fin_raw is None:
                continue

            desc = row[9]
            if desc is None or str(desc).strip() == "":
                continue

            records.append(
                {
                    "Año": year,
                    "Descripción": str(desc).strip(),
                    "Unidad": str(row[10]).strip() if row[10] is not None else "",
                    "Fecha": serialize_date(row[13]),
                    "INICIO (KM)": format_progresiva_excel(inicio_raw),
                    "FIN (KM)": format_progresiva_excel(fin_raw),
                    "_inicio_raw": inicio_raw,
                    "_fin_raw": fin_raw,
                    "METRADO": serialize_metrado(row[16]),
                }
            )
        wb.close()
    return records


def filter_tramo(all_records: list[dict], r_start: int, r_end: int) -> list[dict]:
    filtered = []
    for rec in all_records:
        if strict_within_tramo(rec["_inicio_raw"], rec["_fin_raw"], r_start, r_end):
            filtered.append({k: rec[k] for k in COLUMNS + ["Año"]})
    filtered.sort(key=lambda r: (r["Fecha"] or "", r["Año"], r["INICIO (KM)"] or ""))
    return filtered


header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14, color="1F4E79")
year_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
year_font = Font(bold=True, size=11, color="1F4E79")
thin = Side(style="thin", color="B4B4B4")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")


def autosize_columns(ws):
    widths = [12, 50, 8, 12, 12, 12, 12]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_sheet(ws, tramo_label: str, records: list[dict]):
    ws["A1"] = f"TRAMO km {tramo_label}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        "Filtro estricto: INICIO (KM) y FIN (KM) contenidos en el tramo | "
        f"Total registros: {len(records)}"
    )
    ws.merge_cells("A2:F2")

    current_row = 4
    years = sorted({r["Año"] for r in records})

    if not records:
        ws.cell(current_row, 1, "Sin registros dentro del tramo evaluado.")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        return

    for year in years:
        year_rows = [r for r in records if r["Año"] == year]
        ws.cell(current_row, 1, f"Año {year} ({len(year_rows)} registros)")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        for col in range(1, 7):
            cell = ws.cell(current_row, col)
            cell.fill = year_fill
            cell.font = year_font
            cell.border = border
        current_row += 1

        for col_idx, name in enumerate(COLUMNS, 1):
            cell = ws.cell(current_row, col_idx, name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        current_row += 1

        for rec in year_rows:
            for col_idx, name in enumerate(COLUMNS, 1):
                cell = ws.cell(current_row, col_idx, rec.get(name))
                cell.border = border
                if col_idx == 1:
                    cell.alignment = wrap
            current_row += 1

    ws.freeze_panes = "A5"
    autosize_columns(ws)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    all_records = load_all_records()
    summary = []

    wb_master = Workbook()
    wb_master.remove(wb_master.active)

    for tramo_label, filename, r_start, r_end in TRAMOS:
        records = filter_tramo(all_records, r_start, r_end)
        summary.append((tramo_label, filename, len(records), r_start, r_end))

        payload = {
            "tramo": f"km {tramo_label}",
            "filtro": "estricto: INICIO (KM) y FIN (KM) dentro del tramo",
            "columnas_origen": "Planilla columnas O (INICIO KM) y P (FIN KM)",
            "total_registros": len(records),
            "registros": [{k: r[k] for k in COLUMNS} for r in records],
        }
        with (OUTPUT_DIR / f"{filename}.json").open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        with (OUTPUT_DIR / f"{filename}.csv").open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
            writer.writeheader()
            for r in records:
                writer.writerow({k: r[k] for k in COLUMNS})

        ws = wb_master.create_sheet(title=tramo_label.replace(" ", "")[:31])
        build_sheet(ws, tramo_label, records)

        wb_single = Workbook()
        ws_single = wb_single.active
        ws_single.title = "Reporte"
        build_sheet(ws_single, tramo_label, records)
        wb_single.save(OUTPUT_DIR / f"{filename}.xlsx")

    ws_sum = wb_master.create_sheet(title="Resumen", index=0)
    ws_sum["A1"] = "REPORTES TRAMOS - FILTRO ESTRICTO (INICIO/FIN KM)"
    ws_sum["A1"].font = title_font
    ws_sum.merge_cells("A1:E1")
    headers = ["Tramo (km)", "Registros", "Rango evaluado", "Hoja", "Archivo"]
    for i, h in enumerate(headers, 1):
        c = ws_sum.cell(3, i, h)
        c.fill = header_fill
        c.font = header_font
        c.border = border

    for i, (label, filename, count, r_start, r_end) in enumerate(summary, 4):
        ws_sum.cell(i, 1, label).border = border
        ws_sum.cell(i, 2, count).border = border
        ws_sum.cell(i, 3, f"{format_progresiva_excel(r_start)} a {format_progresiva_excel(r_end)}").border = border
        ws_sum.cell(i, 4, label.replace(" ", "")[:31]).border = border
        ws_sum.cell(i, 5, f"{filename}.xlsx").border = border

    total_row = len(summary) + 5
    ws_sum.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    ws_sum.cell(total_row, 2, sum(c for _, _, c, _, _ in summary)).font = Font(bold=True)

    master_path = OUTPUT_DIR / "reportes_todos_tramos_filtrado_estricto.xlsx"
    wb_master.save(master_path)

    print("RESUMEN FILTRO ESTRICTO")
    print("-" * 60)
    for label, _, count, r_start, r_end in summary:
        print(f"km {label}: {count:4d} registros  [{format_progresiva_excel(r_start)} - {format_progresiva_excel(r_end)}]")
    print(f"TOTAL: {sum(c for _, _, c, _, _ in summary)} registros")
    print(f"\nMaster: {master_path}")


if __name__ == "__main__":
    main()
