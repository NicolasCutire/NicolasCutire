#!/usr/bin/env python3
"""Extract IIRSA Sur base de datos Excel files into structured JSON."""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl

SOURCE_DIR = Path("/workspace/extracted/source/01 BASE DE DATOS")
OUTPUT_DIR = Path("/workspace/extracted/json")
INDEX_PATH = Path("/workspace/extracted/documents_index.json")

SHEET_CONFIG = {
    "Planilla": {
        "header_markers": ("Nº Tramo",),
        "skip_title_patterns": (r"^BASE DE DATOS\s+\d{4}$", r"^CONSOLIDADO DE ACTIVIDADES"),
    },
    "Progresivas": {
        "header_markers": ("Nº Tramo", "Progresiva (Km.) Inicio"),
    },
    "Base": {
        "header_markers": ("Código",),
    },
    "Servicio Secundario": {
        "header_markers": ("Código",),
    },
}


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text.replace("\n", " "))
    return text


def serialize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        return value
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def row_has_data(row: tuple[Any, ...]) -> bool:
    return any(v is not None and str(v).strip() != "" for v in row)


def find_header_row(ws, markers: tuple[str, ...], max_search: int = 40) -> tuple[int | None, list[str]]:
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_search, values_only=True), start=1):
        normalized = [normalize_header(v) for v in row]
        if any(marker in cell for cell in normalized for marker in markers):
            headers = []
            seen: dict[str, int] = {}
            for col_idx, header in enumerate(normalized):
                if not header:
                    header = f"column_{col_idx + 1}"
                if header in seen:
                    seen[header] += 1
                    header = f"{header}_{seen[header]}"
                else:
                    seen[header] = 1
                headers.append(header)
            return row_idx, headers
    return None, []


def extract_title(ws, header_row: int | None) -> str | None:
    search_until = min(header_row or 15, 15)
    for row in ws.iter_rows(max_row=search_until, values_only=True):
        for value in row:
            if value and isinstance(value, str):
                text = value.strip()
                if "CONSOLIDADO" in text.upper() or "BASE DE DATOS" in text.upper():
                    return text
                if "PA214" in text:
                    return text
    return None


def should_skip_row(sheet_name: str, record: dict[str, Any], config: dict) -> bool:
    if not any(v is not None for v in record.values()):
        return True

    patterns = config.get("skip_title_patterns", ())
    for value in record.values():
        if isinstance(value, str):
            for pattern in patterns:
                if re.match(pattern, value.strip(), re.IGNORECASE):
                    return True

    if sheet_name == "Planilla":
        # Skip rows that only contain a year banner in Descripcion
        non_empty = {k: v for k, v in record.items() if v is not None}
        if len(non_empty) == 1 and "Descripcion" in non_empty:
            return True
        # Data rows need at least a service code or metrado
        if record.get("Serv. Secundario") is None and record.get("METRADO") is None:
            if record.get("Nº Tramo") is None:
                return True
    return False


def extract_sheet(wb, sheet_name: str) -> dict[str, Any]:
    ws = wb[sheet_name]
    config = SHEET_CONFIG.get(sheet_name, {"header_markers": ()})
    header_row, headers = find_header_row(ws, config.get("header_markers", ()))

    title = extract_title(ws, header_row)
    records: list[dict[str, Any]] = []

    if header_row and headers:
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row_has_data(row):
                continue
            record = {
                headers[i]: serialize_value(row[i]) if i < len(row) else None
                for i in range(len(headers))
            }
            if should_skip_row(sheet_name, record, config):
                continue
            records.append(record)

    return {
        "sheet_name": sheet_name,
        "title": title,
        "header_row": header_row,
        "columns": headers,
        "row_count": len(records),
        "records": records,
    }


def extract_workbook(xlsx_path: Path) -> dict[str, Any]:
    year_match = re.search(r"(\d{4})", xlsx_path.stem)
    year = int(year_match.group(1)) if year_match else None

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    sheets = {}
    for sheet_name in sheet_names:
        sheets[sheet_name] = extract_sheet(wb, sheet_name)
    wb.close()

    planilla_count = sheets.get("Planilla", {}).get("row_count", 0)
    return {
        "document_id": f"base_datos_{year}",
        "source_file": xlsx_path.name,
        "source_path": str(xlsx_path.relative_to(SOURCE_DIR.parent.parent)),
        "document_type": "excel_database",
        "format": "xlsx",
        "language": "es",
        "year": year,
        "project": "IIRSA Sur - PA214 Cadastro Serviços e Reversões",
        "description": "Consolidado de actividades ejecutadas por frentes de trabajo",
        "sheets": sheet_names,
        "metadata": {
            "total_activity_records": planilla_count,
            "extraction_method": "native_xlsx",
            "is_scanned": False,
        },
        "content": {
            "sheets": sheets,
        },
    }


def build_index(documents: list[dict[str, Any]]) -> dict[str, Any]:
    frentes: set[str] = set()
    rutas: set[str] = set()
    servicios: set[str] = set()
    years: list[int] = []
    total_records = 0

    for doc in documents:
        years.append(doc["year"])
        total_records += doc["metadata"]["total_activity_records"]
        planilla = doc["content"]["sheets"].get("Planilla", {}).get("records", [])
        for rec in planilla:
            if rec.get("Frente IIRSA Sur"):
                frentes.add(str(rec["Frente IIRSA Sur"]))
            if rec.get("Ruta"):
                rutas.add(str(rec["Ruta"]))
            if rec.get("Serv. Secundario"):
                servicios.add(str(rec["Serv. Secundario"]))

    return {
        "index_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_archive": "01 BASE DE DATOS.zip",
        "document_count": len(documents),
        "years_covered": sorted(set(years)),
        "summary": {
            "total_activity_records": total_records,
            "unique_frentes": sorted(frentes),
            "unique_rutas": sorted(rutas),
            "unique_servicios_secundarios_count": len(servicios),
        },
        "documents": [
            {
                "document_id": d["document_id"],
                "year": d["year"],
                "source_file": d["source_file"],
                "json_path": f"extracted/json/{d['document_id']}.json",
                "activity_records": d["metadata"]["total_activity_records"],
                "sheets": d["sheets"],
            }
            for d in sorted(documents, key=lambda x: x["year"] or 0)
        ],
        "schema_reference": {
            "Planilla": {
                "description": "Registro de actividades ejecutadas en campo",
                "key_fields": [
                    "Nº Tramo", "Frente IIRSA Sur", "Ruta", "Inicio Localidad",
                    "Localidad final", "Serv. Secundario", "Descripcion", "Unidad",
                    "UA", "Fecha dd/mm/aa", "INICIO (KM)", "FIN (KM)", "METRADO",
                    "PCI/ITM", "ELEM. SUST.",
                ],
            },
            "Progresivas": {
                "description": "Segmentos de ruta con progresivas kilométricas",
                "key_fields": [
                    "Nº Tramo", "Frente IIRSA Sur", "Ruta", "Inicio Localidad",
                    "Localidad final", "Progresiva (Km.) Inicio", "Progresiva (Km.) Fin",
                    "Long. (Km.)",
                ],
            },
            "Base": {
                "description": "Catálogo de servicios con código UA",
                "key_fields": ["Código", "Descrição", "UMed.", "UA", "Descripcion de la Ua", "Tramo", "Frente"],
            },
            "Servicio Secundario": {
                "description": "Catálogo de códigos de servicio secundario",
                "key_fields": ["Código", "Descrição", "UMed."],
            },
        },
    }


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_files = sorted(SOURCE_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"No xlsx files found in {SOURCE_DIR}", file=sys.stderr)
        return 1

    documents = []
    for xlsx_path in xlsx_files:
        print(f"Processing {xlsx_path.name}...")
        doc = extract_workbook(xlsx_path)
        out_path = OUTPUT_DIR / f"{doc['document_id']}.json"
        with out_path.open("w", encoding="utf-8") as f:
            json.dump(doc, f, ensure_ascii=False, indent=2)
        print(f"  -> {out_path.name}: {doc['metadata']['total_activity_records']} activity records")
        documents.append(doc)

    index = build_index(documents)
    with INDEX_PATH.open("w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)
    print(f"\nIndex written to {INDEX_PATH}")
    print(f"Total documents: {index['document_count']}")
    print(f"Total activity records: {index['summary']['total_activity_records']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
